import csv
from pathlib import Path


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX используя openpyxl.
    """
    # Проверка существования файла
    if not Path(csv_path).exists():
        raise FileNotFoundError(f"Файл {csv_path} не найден")
    
    # Проверка расширения файла
    if not csv_path.lower().endswith('.csv'):
        raise ValueError("Неверный тип файла. Ожидается .csv")
    
    if not xlsx_path.lower().endswith('.xlsx'):
        raise ValueError("Неверный тип файла. Ожидается .xlsx")
    
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Читаем CSV
        with open(csv_path, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            data = list(csv_reader)
        
        # Проверки
        if not data:
            raise ValueError("Пустой CSV файл")
        
        if not data[0]:
            raise ValueError("CSV файл не содержит заголовка")
        
        # Записываем данные
        for row in data:
            ws.append(row)
        
        # Настройка ширины колонок
        for i, column_cells in enumerate(ws.columns, 1):
            length = 8  # минимальная ширина
            for cell in column_cells:
                if cell.value:
                    length = max(length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(i)].width = length + 2
        
        wb.save(xlsx_path)
    
    except ImportError:
        raise ImportError("Для работы функции требуется установить openpyxl: pip install openpyxl")
    except Exception as e:
        raise ValueError(f"Ошибка конвертации: {e}")

csv_to_xlsx('data/samples/cities.csv', 'data/out/output.xlsx')